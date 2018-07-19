from openpyxl import load_workbook
# from all_sentence import *
import pdb
import sys, datetime
import numpy as np
import matplotlib.pyplot as plt
from pylab import mpl
mpl.rcParams['font.sans-serif'] = ['FangSong'] # 指定默认字体
mpl.rcParams['axes.unicode_minus'] = False # 解决保存图像是负号'-'显示为方块的问题

def read_xlsx():
	workbook = load_workbook('WordCup_events_20180615.xlsx', read_only=True)
	worksheet = workbook[workbook.sheetnames[0]]

	all_message = []

	for row in worksheet.rows:
		if row[0].value == "n_ActionID":
			first_line = [x.value for x in row]
			continue
		tmp_dict = {}
		for i in range(2, len(row)):
			tmp_dict[first_line[i]] = row[i].value
		if tmp_dict["n_HomeOrAway"] == '1': home = tmp_dict["c_Team"]
		elif tmp_dict["n_HomeOrAway"] == '-1': away = tmp_dict["c_Team"]
		all_message.append(tmp_dict)

	def get_time(elem):
		return int(elem["n_ActionTime"])
	all_message.sort(key = get_time)
	return all_message, home, away
	# for x in all_message:
	# 	print(x["n_ActionTime"], x["c_Action"])

def count_word(text_list):
	tmp = ''.join(text_list)
	return len(tmp)

def draw_histogram(stats):
	team1 = list(stats.keys())[0]
	team2 = list(stats.keys())[1]
	stats1 = stats[team1]
	stats2 = stats[team2]
	term = np.array(["黄牌", "红牌", "犯规", "越位", "角球", "任意球", "射门", "射正", "扑救", "进球"])
	n = len(term)
	X = np.arange(n)
	Y1 = []
	Y2 = []
	for i in term:
		Y1.append(stats1.get(i,0))
		Y2.append(stats2.get(i,0))
	Y1 = np.array(Y1)
	Y2 = np.array(Y2)
	m = max(max(Y1), max(Y2))
	color1 = []
	color2 = []
	for y1, y2 in zip(Y1, Y2):
		if y1 > y2:
			color1.append("lime")
			color2.append("grey")
		elif y1 < y2:
			color1.append("grey")
			color2.append("lime")
		else:
			color1.append("grey")
			color2.append("grey")
	plt.barh(X, +Y1, color = color1, edgecolor='white')
	plt.barh(X, -Y2, color = color2, edgecolor='white')

	for x,y in zip(X,Y1):
	    plt.text(y+1, x, '%d' % y, ha='center', va= 'center')
	for x,y in zip(X,Y2):
	    plt.text(-(y+1), x, '%d' % y, ha='center', va= 'center')
	for x,y in zip(X,term):
	    plt.text(-(m+6), x, '%s' % y, ha='center', va= 'center')
	plt.text(10, n, team1, ha='center', va= 'center')
	plt.text(-10, n, team2, ha='center', va= 'center')
	plt.xlim(-(m+6),m+1)
	plt.axis('off')
	plt.savefig("{:s}VS{:s}数据统计图.jpg".format(team1, team2))

if __name__ == "__main__":
	all_message, home, away = read_xlsx()
	statistic = {}
	for i, message in enumerate(all_message):
		# pdb.set_trace()	
		if not message["c_Team"] in statistic.keys():
			statistic[message["c_Team"]] = {}
		if not message["c_Action"] in statistic[message["c_Team"]].keys():
			statistic[message["c_Team"]][message["c_Action"]] = 1
		else:
			statistic[message["c_Team"]][message["c_Action"]] += 1
		if message["c_Action"] == "进球":
			if all_message[i-1]["d_ActionDateUTC"] == all_message[i]["d_ActionDateUTC"]:
				if not all_message[i-1]["c_Action"] + "直接破门" in statistic[message["c_Team"]].keys():
					statistic[message["c_Team"]][all_message[i-1]["c_Action"] + "直接破门"] = 1
				else:
					statistic[message["c_Team"]][all_message[i-1]["c_Action"] + "直接破门"] += 1
			else:
				t1 = datetime.datetime.strptime(all_message[i]["d_ActionDateUTC"].split('.')[0], "%Y-%m-%dT%H:%M:%S")
				t2 = datetime.datetime.strptime(all_message[i-1]["d_ActionDateUTC"].split('.')[0], "%Y-%m-%dT%H:%M:%S")
				if (t1 - t2).seconds < 10:
					if not all_message[i-1]["c_Action"] + "直接破门" in statistic[message["c_Team"]].keys():
						statistic[message["c_Team"]][all_message[i-1]["c_Action"] + "间接破门"] = 1
					else:
						statistic[message["c_Team"]][all_message[i-1]["c_Action"] + "间接破门"] += 1
	print("俄罗斯联邦", statistic["俄罗斯联邦"])
	print("沙特阿拉伯", statistic["沙特阿拉伯"])
	game_time = all_message[0]["d_ActionDateUTC"].split("T")[0].split('-')
	game_time = "{:s}年{:s}月{:s}日".format(game_time[0],game_time[1][-1],game_time[2])
	print(game_time)
	draw_histogram(statistic)